from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from datetime import datetime
from openpyxl.styles import Border, Side
import pandas as pd

def parse_start_date(raw: str) -> datetime:
    # Accept YYYY-MM-DD
    raw += "T07:00:00"
    try:
        return datetime.fromisoformat(raw) # ISO datetime strings like 2026-01-15T07:00:00
    except ValueError as exc:
        raise SystemExit(
            "Invalid date format. Use YYYY-MM-DD"
        ) from exc


def normalize_style(style_code: Any) -> str:
    if pd.isna(style_code):
        return ""
    style_code = style_code.strip()
    if len(style_code)==0:
        return ""
    style_code = style_code.upper()
    [style, size] = style_code.split()
    if '2612' in style:
        return style+'-'+size
    if '-' in style:
        style = style.split('-')[0]
    if 'CN' in style:
        style = style[:-2]
    #print(style+'-'+size)
    return style+'-'+size


def format_percentage_columns(ws, df, col_names) -> None:
    for col_name in col_names:
        if col_name in df.columns:
            col_idx = df.columns.get_loc(col_name) + 1  # 1-based for Excel
            col_letter = get_column_letter(col_idx)
            for cell in ws[f"{col_letter}2":f"{col_letter}{ws.max_row}"]:
                cell[0].number_format = "0.00%"


def format_header_bold(ws, font_size: int) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True, size=font_size)


def format_header_fill(ws, hex_color: str) -> None:
    fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
    for cell in ws[1]:
        cell.fill = fill


def apply_row_banding(ws, df, group_col: str, hex_color: str) -> None:
    if df.empty or group_col not in df.columns:
        return

    blue_fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    band_by_value = {}

    for row_idx, value in enumerate(df[group_col].tolist(), start=2):
        if value not in band_by_value:
            band_by_value[value] = len(band_by_value) % 2

        row_fill = blue_fill if band_by_value[value] else white_fill
        for cell in ws[row_idx]:
            cell.fill = row_fill


def auto_fit_columns(ws) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        # Add a small buffer so content doesn't clip
        ws.column_dimensions[col_letter].width = max(max_len + 2, 10)


def apply_table_borders(ws) -> None:
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border